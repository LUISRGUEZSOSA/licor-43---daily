#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
normalize_daily.py
------------------
Normaliza hojas "Daily" heterogéneas de Excel a un CSV largo y uniforme.

Salida (columnas):
  - source_file: nombre del archivo origen
  - sheet: hoja usada (Daily si existe; si no, la primera)
  - section: sección canónica (RESTAURANT, BAR, etc.)
  - metric_label: etiqueta de métrica (texto de fila a la izquierda de las fechas)
  - date: fecha ISO (YYYY-MM-DD) detectada en la cabecera de columnas (forzada al último día del mes)
  - value_raw: texto original de la celda
  - value_num: valor numérico (float), con % escalado a [0..1]
  - is_percent: True si provenía de '%'
  - is_total: True si provenía de una columna TOTAL (a la derecha de fechas)
  - context_left: texto no vacío más cercano a la izquierda de la celda
  - context_top:  texto no vacío más cercano encima de la celda
  - context_left_stack:  hasta 3 etiquetas hacia la izquierda (últimas no vacías, separadas por '|')
  - context_top_stack:   hasta 5 etiquetas hacia arriba (separadas por '|')

Uso:
  python3 normalize_daily.py --out master.csv [--sheet Daily] [--peek] FILE1.xlsx FILE2.xlsx ...

Dependencias:
  pip install pandas openpyxl numpy
"""
import argparse, re, sys, math, unicodedata
from pathlib import Path
from typing import Tuple, Dict, Any
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

# -------------------- Config ----------------------------
SECTION_SYNONYMS = {
    "RESTAURANTE":"RESTAURANT", "RESTAURANT":"RESTAURANT",
    "BAR":"BAR", "BER":"BAR",
    "BODA":"BANQUETING","BENQUETING":"BANQUETING","BANQUETING":"BANQUETING",
    "EMPRESA":"MICE","MICE":"MICE",
    "PARTICULAR":"INDIVIDUALS","INDIVIDUALS":"INDIVIDUALS",
    "TIENDA RESTAURANTE 43":"SHOP","SHOP":"SHOP",
    "WALK IN":"WALKIN","WALKIN":"WALKIN",
    "INTERNO":"EMPLOYEES","EMPLEADOS":"EMPLOYEES","EMPLEADOS:":"EMPLOYEES","EMPLEADO":"EMPLOYEES",
    "EMPLEADOS":"EMPLOYEES",
}

NON_METRIC_LABELS = {"TRUE","FALSE"}

DATE_RE = re.compile(r"^\s*(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})\s*$", re.I)
ISO_RE = re.compile(r"^\d{4}-\d{2}-\d{2}(?:\s+\d{2}:\d{2}:\d{2})?$")
TOTAL_TOKENS = {"TOTAL","TOTAL "}

LABEL_KEYS = {"MES","INICIO DAILY","DIA","DIAS MES","DIA DAILY"}

# --- utilidades para meses (ES/EN y abreviaturas) ---
MONTHS = {
    # español
    "enero":1,"ene":1, "febrero":2,"feb":2, "marzo":3,"mar":3,
    "abril":4,"abr":4, "mayo":5,"may":5, "junio":6,"jun":6,
    "julio":7,"jul":7, "agosto":8,"ago":8, "septiembre":9,"sep":9,"setiembre":9,"set":9,
    "octubre":10,"oct":10, "noviembre":11,"nov":11, "diciembre":12,"dic":12,
    # inglés
    "january":1,"jan":1,"february":2,"feb":2,"march":3,"mar":3,
    "april":4,"apr":4,"may":5,"june":6,"jun":6,"july":7,"jul":7,
    "august":8,"aug":8,"september":9,"sep":9,"october":10,"oct":10,
    "november":11,"nov":11,"december":12,"dec":12,
}

YEAR_ONLY_RE = re.compile(r"^\s*(19|20)\d{2}\s*$")

def month_from_name(s: Any):
    if s is None:
        return None
    key = str(s)
    # normaliza: quita NBSP, dos puntos, puntos, tildes y dobles espacios
    key = key.replace("\u00A0", " ")
    key = key.replace(":", " ").replace(".", " ")
    key = key.strip().lower()
    key = ''.join(c for c in unicodedata.normalize('NFD', key) if unicodedata.category(c) != 'Mn')
    key = re.sub(r"\s+", " ", key)
    return MONTHS.get(key)

def file_month_from_filename(name: str) -> int | None:
    """Detecta mes en el nombre de archivo por texto (ene/feb/...) o número (1..12 o 01..12)."""
    s = name.lower()
    s = ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
    s = s.replace(".xlsx","").replace(".xls","").replace(".csv","")
    s = s.replace("-", " ").replace("_", " ").replace(".", " ")
    tokens = re.split(r"\s+", s)
    # 1) textual
    for t in tokens:
        m = MONTHS.get(t)
        if m: return m
    # 2) número de mes
    for t in tokens:
        if t.isdigit():
            n = int(t)
            if 1 <= n <= 12:
                return n
    return None

def last_day(year: int, month: int) -> pd.Timestamp:
    start = pd.Timestamp(year=year, month=month, day=1)
    return (start + pd.offsets.MonthBegin(1) - pd.offsets.Day(1))

# -------------------- Helpers ----------------------------
def strip_accents(s: str) -> str:
    try:
        return ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
    except Exception:
        return s

def canon(s: Any) -> str:
    if s is None:
        return ""
    s = str(s).strip()
    s = strip_accents(s)
    s = re.sub(r"\s+", " ", s)
    return s

def canon_upper(s: Any) -> str:
    return canon(s).upper()

def is_date_like(x) -> bool:
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return False
    if isinstance(x, (pd.Timestamp, datetime)):
        return True
    if isinstance(x, (np.datetime64,)):
        return True
    s = str(x).strip()
    if not s:
        return False
    try:
        d = pd.to_datetime(s, dayfirst=True, errors="coerce")
        return pd.notna(d) and (1900 <= d.year <= 2100)
    except Exception:
        return False

def is_month_or_date_cell(x) -> bool:
    """True si la celda es una fecha parseable o un nombre de mes."""
    return is_date_like(x) or (month_from_name(x) is not None)

def parse_date(s):
    if s is None:
        return pd.NaT
    if isinstance(s, (pd.Timestamp, datetime)):
        return pd.to_datetime(s).normalize()
    if isinstance(s, (np.datetime64,)):
        return pd.to_datetime(s).normalize()
    s = str(s).strip()
    if not s:
        return pd.NaT
    if ISO_RE.match(s):
        d = pd.to_datetime(s, errors="coerce")
    else:
        d = pd.to_datetime(s, dayfirst=True, errors="coerce")
    if pd.isna(d):
        return pd.NaT
    return pd.to_datetime(d).normalize()

def parse_number(x) -> Tuple[Any, bool]:
    """Devuelve (valor_float|None, is_percent). Escala '%' a [0..1]. Acepta coma decimal y notación científica."""
    if x is None:
        return None, False
    if isinstance(x, float) and math.isnan(x):
        return None, False
    if isinstance(x, (int, float)) and not (isinstance(x, float) and (math.isinf(x) or math.isnan(x))):
        return float(x), False

    s = str(x).strip()
    if not s or s.lower() in {"nan","none","-"}:
        return None, False

    is_pct = s.endswith("%")
    if is_pct:
        s = s[:-1].strip()

    s = s.replace("\u00A0","").replace(" ","").replace("€","").replace("$","").replace("£","")
    s = s.replace(",", ".")  # coma decimal -> punto

    try:
        val = float(s)
    except ValueError:
        return None, is_pct

    if is_pct:
        val /= 100.0
    return val, is_pct

def expand_merged(ws, grid):
    # Propaga valores en celdas combinadas
    for rng in ws.merged_cells.ranges:
        min_row, min_col, max_row, max_col = rng.min_row, rng.min_col, rng.max_row, rng.max_col
        val = grid[min_row-1][min_col-1]
        for r in range(min_row-1, max_row):
            for c in range(min_col-1, max_col):
                grid[r][c] = val
    return grid

def find_date_header_row(grid) -> Tuple[int,int,int,int]:
    """
    Busca la fila con más 'celdas-fecha/mes'. Devuelve (row_idx, first_col, last_col, total_col).
    - total_col: índice de columna que contenga "TOTAL" (a la derecha de fechas), o -1 si no hay.
    """
    best = (-1,-1,-1,-1)
    for r, row in enumerate(grid):
        date_cols = [c for c, v in enumerate(row) if is_month_or_date_cell(v)]
        if len(date_cols) >= 3:
            first_c, last_c = min(date_cols), max(date_cols)
            total_c = -1
            for c in range(last_c+1, len(row)):
                v = row[c]
                if isinstance(v, str) and canon_upper(v) in TOTAL_TOKENS:
                    total_c = c
                    break
            return (r, first_c, last_c, total_c)
    return best

def nearest_left(grid, r, c, max_depth=3) -> Tuple[str,str]:
    vals = []
    cc = c-1
    while cc >= 0 and len(vals) < max_depth:
        txt = canon(grid[r][cc])
        if txt != "":
            vals.append(txt)
        cc -= 1
    return (vals[0] if vals else "", "|".join(vals))

def nearest_top(grid, r, c, max_depth=5) -> Tuple[str,str]:
    vals = []
    rr = r-1
    while rr >= 0 and len(vals) < max_depth:
        txt = canon(grid[rr][c])
        if txt != "":
            vals.append(txt)
        rr -= 1
    return (vals[0] if vals else "", "|".join(vals))

def detect_meta(grid) -> Dict[str, Any]:
    """Busca pares etiqueta-valor típicos (Mes, Inicio Daily, Dia Daily, etc.)."""
    meta = {}
    for r, row in enumerate(grid):
        for c in range(len(row)-1):
            key = canon_upper(row[c])
            if key in LABEL_KEYS:
                meta[key] = canon(row[c+1])
    # parse fechas
    if "INICIO DAILY" in meta:
        d = parse_date(meta["INICIO DAILY"])
        meta["INICIO DAILY"] = str(d.date()) if pd.notna(d) else meta["INICIO DAILY"]
    if "DIA DAILY" in meta:
        d = parse_date(meta["DIA DAILY"])
        meta["DIA DAILY"] = str(d.date()) if pd.notna(d) else meta["DIA DAILY"]
    return meta

def infer_month_last_day(source_file: str, meta_prefixed: dict, col2date: dict, file_month: int | None):
    """Devuelve el último día del mes referente del archivo (date)."""
    month_ref = None

    # 0) Si el nombre del archivo tiene mes textual/numérico, úsalo directamente con un año razonable.
    if file_month:
        # año: intenta meta primero
        start_s = None
        for k in ["INICIO DAILY", "DIA DAILY", "MES"]:
            k2 = f"meta_{k.replace(' ','_').lower()}"
            if k2 in meta_prefixed:
                start_s = meta_prefixed[k2]
                break
        base_year = None
        if start_s:
            dt = pd.to_datetime(start_s, errors="coerce")
            if pd.notna(dt):
                base_year = int(dt.year)
        if base_year is None:
            m = re.search(r"(\d{4})", source_file)
            if m:
                base_year = int(m.group(1))
        if base_year is None:
            # último recurso: mayor año visible en col2date
            years = [pd.to_datetime(v).year for v in col2date.values()] if col2date else []
            base_year = max(years) if years else pd.Timestamp.today().year
        return (last_day(base_year, file_month)).date()

    # 1) meta explícita
    start_s = None
    for k in ["INICIO DAILY", "DIA DAILY", "MES"]:
        k2 = f"meta_{k.replace(' ','_').lower()}"
        if k2 in meta_prefixed:
            start_s = meta_prefixed[k2]
            break
    if start_s:
        dt = pd.to_datetime(start_s, errors="coerce")
        if pd.notna(dt):
            month_ref = pd.Timestamp(year=dt.year, month=dt.month, day=1)

    # 2) filename YYYYMM
    if month_ref is None:
        m = re.search(r"(\d{4})(\d{2})", source_file)
        if m:
            y, mth = int(m.group(1)), int(m.group(2))
            try:
                month_ref = pd.Timestamp(year=y, month=mth, day=1)
            except Exception:
                month_ref = None

    # 3) fallback: max fecha en col2date
    if month_ref is None and col2date:
        try:
            mx = max([pd.to_datetime(v) for v in col2date.values() if pd.notna(v)])
            month_ref = pd.Timestamp(year=mx.year, month=mx.month, day=1)
        except Exception:
            month_ref = None

    if month_ref is None:
        month_ref = pd.Timestamp.today().normalize().replace(day=1)

    next_month = (month_ref + pd.offsets.MonthBegin(1))
    return (next_month - pd.offsets.Day(1)).date()

def normalize_one(path: Path, sheet_name: str = "Daily") -> pd.DataFrame:
    wb = load_workbook(path, data_only=True)
    sheet = sheet_name if sheet_name in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet]
    max_row, max_col = ws.max_row, ws.max_column
    grid = [[ws.cell(row=r, column=c).value for c in range(1, max_col+1)] for r in range(1, max_row+1)]
    grid = expand_merged(ws, grid)

    # detectar cabecera de fechas/meses
    hdr_row, first_c, last_c, total_c = find_date_header_row(grid)
    if hdr_row < 0:
        raise RuntimeError(f"No encontré fila de fechas/meses en {path.name} (sheet {sheet})")

    # meta antes de mapear fechas
    meta = detect_meta(grid)

    # Año base (para construir fechas cuando solo sabemos el mes)
    base_year = None
    for k in ("INICIO DAILY", "DIA DAILY"):
        if k in meta:
            d0 = pd.to_datetime(meta[k], errors="coerce", dayfirst=True)
            if pd.notna(d0):
                base_year = int(d0.year)
                break
    if base_year is None:
        m = re.search(r"(\d{4})", path.name)
        if m:
            base_year = int(m.group(1))
    if base_year is None:
        base_year = pd.Timestamp.today().year

    # Mes deducido del nombre de archivo (p.ej. "02_febrero.xlsx", "03 marzo.xlsx")
    file_month = file_month_from_filename(path.name)

    # mapa col->fecha (forzando último día del mes)
    col2date: Dict[int, pd.Timestamp] = {}
    for c in range(first_c, last_c+1):
        cell = grid[hdr_row][c]
        sraw = "" if cell is None else str(cell).strip()
        d = parse_date(cell)  # intento parseo estándar (fechas tipo 01/02/2025)

        if sraw and YEAR_ONLY_RE.match(sraw):
            # header es solo "2025" o "2024" → usa ese año + MES DEL ARCHIVO
            y = int(sraw)
            mnum = file_month if file_month else pd.Timestamp.today().month
            d = last_day(y, mnum)

        elif pd.isna(d):
            mnum = month_from_name(cell)  # p.ej. "febrero" -> 2
            if mnum:
                d = last_day(base_year, mnum)
            elif file_month:
                # fallback: usa el mes del archivo si la celda no da fecha ni mes
                d = last_day(base_year, file_month)
        else:
            # ya hay fecha -> llevarla al último día de su mes
            d = last_day(d.year, d.month)

        if pd.notna(d):
            col2date[c] = d

    # compute last day de referencia para 'scenario'
    meta_prefixed = {f"meta_{k.replace(' ','_').lower()}": v for k, v in meta.items()}
    last_day_for_file = infer_month_last_day(path.name, meta_prefixed, col2date, file_month)

    current_section = None
    rows = []

    # recorre debajo del header
    for r in range(hdr_row+1, len(grid)):
        row = grid[r]

        # etiqueta más a la izquierda antes de las fechas
        label = None
        for c in range(0, first_c):
            v = row[c]
            if v is not None:
                s = canon(v)
                if s != "":
                    label = s  # última no vacía
        metric_label = None
        if label:
            up = canon_upper(label)
            if up in SECTION_SYNONYMS:
                current_section = SECTION_SYNONYMS[up]
            elif up in NON_METRIC_LABELS:
                metric_label = None
            else:
                metric_label = label

        # celdas diarias
        for c, d in col2date.items():
            v = row[c]
            if v is None:
                continue
            s = canon(v)
            if s == "" or is_month_or_date_cell(v):  # evitar re-leer encabezados
                continue
            val_num, is_pct = parse_number(s)
            if val_num is None:
                continue

            left1, left_stack = nearest_left(grid, r, c)
            top1, top_stack   = nearest_top(grid, r, c)

            scen = "TODAY" if pd.to_datetime(d).date() == last_day_for_file else "MONTH_TO_DATE"
            rows.append({
                "source_file": path.name,
                "sheet": sheet,
                "section": current_section,
                "metric_label": metric_label,
                "date": str(pd.to_datetime(d).date()),
                "value_raw": s,
                "value_num": val_num,
                "is_percent": is_pct,
                "is_total": False,
                "scenario": scen,
                "context_left": left1,
                "context_top": top1,
                "context_left_stack": left_stack,
                "context_top_stack": top_stack,
                **{f"meta_{k.replace(' ','_').lower()}": v for k, v in meta.items()},
            })

        # TOTAL si existe
        if 0 <= total_c < len(row):
            tv = row[total_c]
            if tv is not None:
                s = canon(tv)
                if s != "":
                    val_num, is_pct = parse_number(s)
                    if val_num is not None:
                        left1, left_stack = nearest_left(grid, r, total_c)
                        top1, top_stack   = nearest_top(grid, r, total_c)
                        rows.append({
                            "source_file": path.name,
                            "sheet": sheet,
                            "section": current_section,
                            "metric_label": metric_label,
                            "date": "",  # total sin fecha
                            "value_raw": s,
                            "value_num": val_num,
                            "is_percent": is_pct,
                            "is_total": True,
                            "scenario": "TOTAL",
                            "context_left": left1,
                            "context_top": top1,
                            "context_left_stack": left_stack,
                            "context_top_stack": top_stack,
                            **{f"meta_{k.replace(' ','_').lower()}": v for k, v in meta.items()},
                        })

    out = pd.DataFrame.from_records(rows)
    # ordenar columnas
    base_cols = [
        "source_file","sheet","section","metric_label","date",
        "value_raw","value_num","is_percent","is_total",
        "context_left","context_top","context_left_stack","context_top_stack","scenario",
    ]
    meta_cols = sorted([c for c in out.columns if c.startswith("meta_")])
    out = out.reindex(columns=base_cols + meta_cols)
    return out

def main():
    import glob
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default="master_daily_normalized.csv", help="Ruta del CSV maestro (formato largo)")
    ap.add_argument("--sheet", default="Daily", help="Nombre de la hoja (default: Daily)")
    ap.add_argument("--peek", action="store_true", help="Guarda _peek_*.csv (primeras 30 filas por archivo)")
    ap.add_argument("files", nargs="*", help="Rutas a .xlsx/.xls (si se omite, usa *.xlsx del directorio actual)")
    args = ap.parse_args()

    file_list = args.files if args.files else sorted(glob.glob("*.xlsx"))
    if not file_list:
        print("No se encontraron Excel (*.xlsx).")
        return 1

    frames = []
    for f in file_list:
        path = Path(f)
        try:
            df = normalize_one(path, sheet_name=args.sheet)
            if args.peek:
                df.head(30).to_csv(f"_peek_{path.stem}.csv", index=False)
            frames.append(df)
            print(f"[OK] {path.name}: {len(df)} filas")
        except Exception as e:
            print(f"[ERR] {path.name}: {e}")

    master = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
    master.to_csv(args.out, index=False, encoding="utf-8")
    by_file = master.groupby("source_file")["value_raw"].count().to_dict() if not master.empty else {}
    print("OK. Filas por archivo:", by_file)
    print("Salida:", args.out)
    return 0

if __name__ == "__main__":
    sys.exit(main())
